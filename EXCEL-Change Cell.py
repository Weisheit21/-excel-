import time
import openpyxl
from openpyxl.reader.excel import load_workbook
import os
import os.path

def open_xlsx(filename):
    # 加载Excel数据，处理数据
    wb = load_workbook(filename)    # 打开excel文件
    ws = wb.worksheets[0]           # 获取
    # 修改 
    ws['F5'] = '2020/2/24'
    ws['F4'] = '周天奇'
    ws['C5'] = '2020/2/20'
    ws['C4'] = '李巧明'
    ws['F3'] = '2020年度'
    ws.column_dimensions['B'].width = 8     #给被审计单位留出空间
    wb.save(filename)                       # 保存修改后的excel

def openPath(path=""):
    fileList = os.listdir(path)   # 获取path目录下所有文件
    for filename in fileList:
        pathTmp = os.path.join(path,filename)   # 获取path与filename组合后的路径
        if os.path.isdir(pathTmp):   # 如果是目录
            openPath(pathTmp)        # 则递归查找

        else: # 文件
            (file, ext) = os.path.splitext(filename)
            if ext == ".xlsx" or ext == ".XLSx":
                print(pathTmp + 'is being processed!')
                open_xlsx(pathTmp)

def main():

    time_start = time.time()

    path=r'C:\Users\11647\Desktop\2019 惠金所 底稿【4】'
    openPath(path)

    time_end = time.time()
    print('😎😎本次运行花费时间:',round(time_end-time_start,2),'秒')

if __name__ == '__main__':
    main()


